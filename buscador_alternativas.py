#!/usr/bin/env python3
"""
AGENTE BUSCADOR DE ALTERNATIVAS DE CONTACTO
Para contribuyentes inubicables (sin mail ni teléfono)
Busca en la base de datos: mismo apellido + misma dirección
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

def normalizar_apellido(nombre):
    """Extrae y normaliza el apellido"""
    if pd.isna(nombre):
        return None
    partes = str(nombre).strip().split()
    return partes[0].upper() if partes else None

def normalizar_direccion(dir_str):
    """Normaliza la dirección para búsqueda"""
    if pd.isna(dir_str):
        return None
    # Convertir a mayúsculas y eliminar espacios múltiples
    return ' '.join(str(dir_str).strip().upper().split())

def buscar_alternativas(df, idx_deudor, row_deudor):
    """Busca alternativas de contacto para un deudor inubicable"""

    apellido = row_deudor['APELLIDO']
    direccion = row_deudor['DIRECCION_NORM']

    # PRIORIDAD 1: Mismo apellido + Misma dirección
    grupo1 = df[
        (df['APELLIDO'] == apellido) &
        (df['DIRECCION_NORM'] == direccion) &
        (df.index != idx_deudor) &
        ((df['MAIL PARA ENVIO'].notna()) | (df['CELULAR PARA ENVIO'].notna()))
    ]

    # PRIORIDAD 2: Mismo apellido (cualquier dirección)
    grupo2 = df[
        (df['APELLIDO'] == apellido) &
        (df.index != idx_deudor) &
        ((df['MAIL PARA ENVIO'].notna()) | (df['CELULAR PARA ENVIO'].notna()))
    ]

    # PRIORIDAD 3: Misma dirección (cualquier apellido)
    grupo3 = df[
        (df['DIRECCION_NORM'] == direccion) &
        (df.index != idx_deudor) &
        ((df['MAIL PARA ENVIO'].notna()) | (df['CELULAR PARA ENVIO'].notna()))
    ]

    return {
        'grupo1': grupo1.head(2),  # Mejor alternativa
        'grupo2': grupo2.head(2),  # Alternativa media
        'grupo3': grupo3.head(2)   # Alternativa débil
    }

def main():
    print("\n" + "="*80)
    print("AGENTE BUSCADOR DE ALTERNATIVAS DE CONTACTO")
    print("="*80)

    # Cargar base de datos
    print("\n📂 Cargando base de datos...")
    df = pd.read_excel("../uploads/B - DETALLE DE TODOS LOS OBJETOS.xlsx")

    # Contar objetos por contribuyente (el Excel tiene una fila por objeto)
    objetos_por_cuit = df.groupby('CUIT FINAL').size()

    # Preparar datos
    df_trabajo = df[['CUIT FINAL', 'DENOMINACION FINAL', 'DOMICILIO_CATASTRO',
                     'MAIL PARA ENVIO', 'CELULAR PARA ENVIO',
                     'Deuda Corriente Actualizada']].copy()

    # Normalizar
    df_trabajo['APELLIDO'] = df_trabajo['DENOMINACION FINAL'].apply(normalizar_apellido)
    df_trabajo['DIRECCION_NORM'] = df_trabajo['DOMICILIO_CATASTRO'].apply(normalizar_direccion)
    df_trabajo['SIN_CONTACTO'] = (df_trabajo['MAIL PARA ENVIO'].isna()) & (df_trabajo['CELULAR PARA ENVIO'].isna())

    # Filtrar inubicables
    inubicables = df_trabajo[df_trabajo['SIN_CONTACTO']].copy()

    print(f"\n📊 ESTADÍSTICAS:")
    print(f"   Total contribuyentes: {len(df_trabajo):,}")
    print(f"   Inubicables (sin contacto): {len(inubicables):,}")
    print(f"   Porcentaje: {len(inubicables)/len(df_trabajo)*100:.1f}%")
    print(f"   Deuda total inubicables: ${inubicables['Deuda Corriente Actualizada'].sum():,.2f}")

    # Buscar alternativas
    print(f"\n🔍 Buscando alternativas...")

    resultados = []
    encontrados_grupo1 = 0
    encontrados_grupo2 = 0
    encontrados_grupo3 = 0

    for idx, row in inubicables.iterrows():
        alternativas = buscar_alternativas(df_trabajo, idx, row)

        # Determinar mejor alternativa
        mejor = None
        prioridad = None

        if len(alternativas['grupo1']) > 0:
            mejor = alternativas['grupo1'].iloc[0]
            prioridad = 1
            encontrados_grupo1 += 1
        elif len(alternativas['grupo2']) > 0:
            mejor = alternativas['grupo2'].iloc[0]
            prioridad = 2
            encontrados_grupo2 += 1
        elif len(alternativas['grupo3']) > 0:
            mejor = alternativas['grupo3'].iloc[0]
            prioridad = 3
            encontrados_grupo3 += 1

        if mejor is not None:
            # Describir el tipo de relación
            if prioridad == 1:
                tipo = "👤 MISMO APELLIDO + MISMA DIRECCIÓN"
            elif prioridad == 2:
                tipo = "👥 MISMO APELLIDO"
            else:
                tipo = "📍 MISMA DIRECCIÓN"

            # Obtener contacto
            telefono = mejor['CELULAR PARA ENVIO'] if pd.notna(mejor['CELULAR PARA ENVIO']) else ''
            email = mejor['MAIL PARA ENVIO'] if pd.notna(mejor['MAIL PARA ENVIO']) else ''

            resultados.append({
                'CUIT': int(row['CUIT FINAL']),
                'DEUDOR': row['DENOMINACION FINAL'],
                'DIRECCIÓN': row['DOMICILIO_CATASTRO'],
                'DEUDA': row['Deuda Corriente Actualizada'],
                'CANT_OBJETOS': int(objetos_por_cuit.get(row['CUIT FINAL'], 1)),
                'TIPO_BÚSQUEDA': tipo,
                'CONTACTO_NOMBRE': mejor['DENOMINACION FINAL'],
                'CONTACTO_TELÉFONO': telefono,
                'CONTACTO_EMAIL': email,
                'PRIORIDAD': prioridad
            })

    # Crear DataFrame de resultados
    df_resultados = pd.DataFrame(resultados)
    df_resultados = df_resultados.sort_values('PRIORIDAD')

    print(f"\n✅ RESULTADOS:")
    print(f"   Alternativas encontradas: {len(df_resultados):,}")
    print(f"   - Prioridad 1 (Mismo apellido + dirección): {encontrados_grupo1:,}")
    print(f"   - Prioridad 2 (Mismo apellido): {encontrados_grupo2:,}")
    print(f"   - Prioridad 3 (Misma dirección): {encontrados_grupo3:,}")
    print(f"   Sin alternativa encontrada: {len(inubicables) - len(df_resultados):,}")

    # Guardar Excel
    archivo = "INUBICABLES_CON_ALTERNATIVAS.xlsx"
    print(f"\n💾 Guardando en {archivo}...")

    df_resultados.to_excel(archivo, index=False, sheet_name='Alternativas')

    # Formatear Excel
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # Colorear encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ancho de columnas
    ws.column_dimensions['A'].width = 12  # CUIT
    ws.column_dimensions['B'].width = 22  # DEUDOR
    ws.column_dimensions['C'].width = 28  # DIRECCIÓN
    ws.column_dimensions['D'].width = 14  # DEUDA
    ws.column_dimensions['E'].width = 14  # CANT_OBJETOS
    ws.column_dimensions['F'].width = 35  # TIPO_BÚSQUEDA
    ws.column_dimensions['G'].width = 22  # CONTACTO_NOMBRE
    ws.column_dimensions['H'].width = 16  # CONTACTO_TELÉFONO
    ws.column_dimensions['I'].width = 24  # CONTACTO_EMAIL
    ws.column_dimensions['J'].width = 10  # PRIORIDAD

    # Alineación
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(archivo)

    # Mostrar ejemplos
    print(f"\n📋 EJEMPLOS DE RESULTADOS:")
    print(df_resultados[['CUIT', 'DEUDOR', 'TIPO_BÚSQUEDA', 'CONTACTO_NOMBRE',
                         'CONTACTO_TELÉFONO']].head(10).to_string(index=False))

    print(f"\n" + "="*80)
    print(f"✨ Proceso completado - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()
